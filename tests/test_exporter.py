# tests/test_exporter.py
import pytest
import os
from openpyxl import load_workbook
from output.exporter import export_xlsx, export_csv

def test_export_xlsx_creates_4_sheets(tmp_path):
    prospects = [
        {"company_name": "Acme", "state": "TX", "score": 80, "tier": "HOT",
         "status": "NEW", "email_verified": "valid", "first_seen": "2026-03-24",
         "vertical": "food", "contact_email": "j@acme.com", "source_channel": "web_search",
         "city": "", "address": "", "zip_code": "", "phone": "", "website": "acme.com",
         "estimated_employees": 100, "estimated_revenue": 10000000, "product_keywords": "",
         "compliance_signals": "", "contact_name": "John", "contact_title": "VP Logistics",
         "contact_source": "apollo", "email_confidence": 91, "registration_id": "",
         "import_products": "", "notes": "", "score_breakdown": "test", "normalized_name": "acme",
         "domain": "acme.com", "last_seen": "2026-03-24", "run_count": 1,
         "status_updated": "", "status_notes": ""},
        {"company_name": "Beta", "state": "CA", "score": 30, "tier": "NURTURE",
         "status": "NEW", "email_verified": "", "first_seen": "2026-03-24",
         "vertical": "pharma", "contact_email": "", "source_channel": "apollo",
         "city": "", "address": "", "zip_code": "", "phone": "", "website": "beta.com",
         "estimated_employees": None, "estimated_revenue": None, "product_keywords": "",
         "compliance_signals": "", "contact_name": "", "contact_title": "",
         "contact_source": "", "email_confidence": None, "registration_id": "",
         "import_products": "", "notes": "", "score_breakdown": "test", "normalized_name": "beta",
         "domain": "beta.com", "last_seen": "2026-03-24", "run_count": 1,
         "status_updated": "", "status_notes": ""}
    ]
    run_history = [{"run_date": "2026-03-24", "new_count": 2, "dedup_count": 2,
                    "states_used": "TX", "verticals_used": "food", "channels_used": "web_search",
                    "raw_count": 5, "updated_count": 0, "hot_count": 1, "warm_count": 0,
                    "nurture_count": 1, "park_count": 0, "avg_score": 55.0,
                    "duration_seconds": 120, "serpapi_credits": 50, "apollo_credits": 100,
                    "hunter_search_credits": 10, "hunter_verify_credits": 0}]
    pipeline_stats = {"status_counts": {"NEW": 2}, "tier_counts": {"HOT": 1, "NURTURE": 1}}

    path = str(tmp_path / "test.xlsx")
    export_xlsx(path, prospects, run_history, pipeline_stats, run_date="2026-03-24")

    wb = load_workbook(path)
    assert len(wb.sheetnames) == 4
    assert wb.sheetnames[0] == "New This Run"
    assert wb.sheetnames[1] == "Full Prospects"
    assert wb.sheetnames[2] == "Pipeline Dashboard"
    assert wb.sheetnames[3] == "Run Log"
    ws = wb["New This Run"]
    assert ws.max_row >= 2

def test_export_csv_creates_file(tmp_path):
    prospects = [
        {"company_name": "Acme", "state": "TX", "score": 80, "tier": "HOT",
         "website": "acme.com", "vertical": "food", "source_channel": "web_search"}
    ]
    path = str(tmp_path / "test.csv")
    export_csv(path, prospects)
    assert os.path.exists(path)
    with open(path) as f:
        lines = f.readlines()
    assert len(lines) == 2
