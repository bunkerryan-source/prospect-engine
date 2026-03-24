"""
Excel and CSV export for prospect data.
Produces professionally formatted, shareable Excel files for the sales team.
"""

import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Color constants (hex, no #) ───────────────────────────────────────────────
NAVY = "003366"
WHITE = "FFFFFF"
ALT_ROW = "F2F2F2"

TIER_COLORS = {
    "HOT": "FF4444",
    "WARM": "FF8C00",
    "NURTURE": "FFD700",
    "PARK": "C0C0C0",
}

STATUS_COLORS = {
    "NEW": "4A90D9",
    "CONTACTED": "FF8C00",
    "ENGAGED": "28A745",
    "WON": "1B5E20",
    "LOST": "DC3545",
    "PARKED": "C0C0C0",
}

EMAIL_VERIFIED_COLORS = {
    "valid": "28A745",
    "invalid": "DC3545",
    "accept_all": "FFD700",
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _bold_font(color=None):
    return Font(bold=True, color=color or "000000")


def _header_font():
    return Font(bold=True, color=WHITE)


def _center():
    return Alignment(horizontal="center", vertical="center")


def _apply_header_row(ws, headers):
    """Write navy header row with white bold text."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _fill(NAVY)
        cell.font = _header_font()
        cell.alignment = _center()


def _auto_column_widths(ws):
    """Set column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 50)


def _write_prospects_sheet(ws, prospects):
    """Write a full prospects sheet with formatting (used for Sheet 1 and 2)."""
    columns = [
        "company_name", "city", "state", "phone", "website", "vertical",
        "source_channel", "estimated_employees", "estimated_revenue",
        "contact_name", "contact_title", "contact_email", "email_verified",
        "score", "tier", "status", "first_seen", "last_seen", "notes",
    ]
    headers = [c.replace("_", " ").title() for c in columns]

    _apply_header_row(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    col_index = {col: idx + 1 for idx, col in enumerate(columns)}

    sorted_prospects = sorted(prospects, key=lambda p: (p.get("score") or 0), reverse=True)

    for row_idx, prospect in enumerate(sorted_prospects, start=2):
        is_alt = (row_idx % 2 == 0)
        row_bg = ALT_ROW if is_alt else None

        for col_name in columns:
            col_num = col_index[col_name]
            value = prospect.get(col_name)
            cell = ws.cell(row=row_idx, column=col_num, value=value)

            # Default row shading
            if row_bg:
                cell.fill = _fill(row_bg)

            # Color-coded columns
            if col_name == "tier" and value in TIER_COLORS:
                cell.fill = _fill(TIER_COLORS[value])
                cell.font = _bold_font()
                cell.alignment = _center()
            elif col_name == "status" and value in STATUS_COLORS:
                cell.fill = _fill(STATUS_COLORS[value])
                cell.font = Font(bold=True, color=WHITE)
                cell.alignment = _center()
            elif col_name == "email_verified" and value in EMAIL_VERIFIED_COLORS:
                cell.fill = _fill(EMAIL_VERIFIED_COLORS[value])
                cell.font = Font(bold=True, color=WHITE)
                cell.alignment = _center()
            elif col_name == "score":
                cell.number_format = "0"

    _auto_column_widths(ws)


# ── Pipeline Dashboard helpers ────────────────────────────────────────────────

def _write_dashboard_title(ws, run_date):
    ws["A1"] = "Pipeline Dashboard"
    ws["A1"].font = Font(bold=True, size=16, color=NAVY)
    ws["A2"] = f"Run Date: {run_date}"
    ws["A2"].font = Font(italic=True, color="555555")


def _write_section_header(ws, row, label):
    cell = ws.cell(row=row, column=1, value=label)
    cell.fill = _fill(NAVY)
    cell.font = _header_font()
    ws.cell(row=row, column=2).fill = _fill(NAVY)


def _write_kv_table(ws, start_row, header_label, data_pairs):
    """Write a two-column key/value table with a section header above."""
    _write_section_header(ws, start_row, header_label)
    for offset, (key, value) in enumerate(data_pairs, start=1):
        r = start_row + offset
        ws.cell(row=r, column=1, value=key)
        ws.cell(row=r, column=2, value=value)


# ── Public API ────────────────────────────────────────────────────────────────

def export_xlsx(path, prospects, run_history, pipeline_stats, run_date):
    """
    Export data to a formatted .xlsx workbook with 4 sheets:
      1. New This Run
      2. Full Prospects
      3. Pipeline Dashboard
      4. Run Log
    """
    wb = Workbook()

    # ── Sheet 1: New This Run ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "New This Run"
    new_prospects = [p for p in prospects if p.get("first_seen") == run_date]
    _write_prospects_sheet(ws1, new_prospects)

    # ── Sheet 2: Full Prospects ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Full Prospects")
    _write_prospects_sheet(ws2, prospects)

    # ── Sheet 3: Pipeline Dashboard ───────────────────────────────────────────
    ws3 = wb.create_sheet("Pipeline Dashboard")
    _write_dashboard_title(ws3, run_date)

    # Status breakdown (rows 4–12)
    status_counts = pipeline_stats.get("status_counts", {})
    status_pairs = [(s, c) for s, c in status_counts.items()]
    _write_kv_table(ws3, 4, "Status Breakdown", status_pairs[:8])

    # Tier distribution (rows 14–19)
    tier_counts = pipeline_stats.get("tier_counts", {})
    tier_pairs = [(t, c) for t, c in tier_counts.items()]
    _write_kv_table(ws3, 14, "Tier Distribution", tier_pairs[:5])

    # Run summary (rows 21–23)
    new_count = len([p for p in prospects if p.get("first_seen") == run_date])
    returning_count = len(prospects) - new_count
    _write_section_header(ws3, 21, "Run Summary")
    ws3.cell(row=22, column=1, value="New")
    ws3.cell(row=22, column=2, value=new_count)
    ws3.cell(row=23, column=1, value="Returning")
    ws3.cell(row=23, column=2, value=returning_count)
    ws3.cell(row=24, column=1, value="Total")
    ws3.cell(row=24, column=2, value=len(prospects))

    # Top 10 new prospects (rows 25–36)
    top_new = sorted(
        [p for p in prospects if p.get("first_seen") == run_date],
        key=lambda p: (p.get("score") or 0),
        reverse=True,
    )[:10]
    top_headers = ["Company", "State", "Vertical", "Score", "Tier", "Contact Email"]
    _write_section_header(ws3, 25, "Top New Prospects")
    for col_idx, h in enumerate(top_headers, start=1):
        cell = ws3.cell(row=25, column=col_idx, value=h)
        cell.fill = _fill(NAVY)
        cell.font = _header_font()
    for row_offset, p in enumerate(top_new, start=1):
        r = 25 + row_offset
        ws3.cell(row=r, column=1, value=p.get("company_name"))
        ws3.cell(row=r, column=2, value=p.get("state"))
        ws3.cell(row=r, column=3, value=p.get("vertical"))
        ws3.cell(row=r, column=4, value=p.get("score"))
        ws3.cell(row=r, column=5, value=p.get("tier"))
        ws3.cell(row=r, column=6, value=p.get("contact_email"))

    _auto_column_widths(ws3)

    # ── Sheet 4: Run Log ──────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Run Log")
    if run_history:
        log_headers = list(run_history[0].keys())
        _apply_header_row(ws4, log_headers)
        ws4.freeze_panes = "A2"
        for row_idx, entry in enumerate(run_history, start=2):
            is_alt = (row_idx % 2 == 0)
            for col_idx, key in enumerate(log_headers, start=1):
                cell = ws4.cell(row=row_idx, column=col_idx, value=entry.get(key))
                if is_alt:
                    cell.fill = _fill(ALT_ROW)
        _auto_column_widths(ws4)

    wb.save(path)


def export_csv(path, prospects):
    """
    Export all prospects to a CSV file.
    Header row = dict keys; data rows sorted by score descending.
    """
    if not prospects:
        with open(path, "w", newline="") as f:
            pass
        return

    sorted_prospects = sorted(prospects, key=lambda p: (p.get("score") or 0), reverse=True)
    fieldnames = list(prospects[0].keys())

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(sorted_prospects)
